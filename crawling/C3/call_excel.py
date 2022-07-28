import openpyxl

fpath = r'C:\python_crawling\C3\참가자_data.xlsx'

# 엑셀 불러오기
wb = openpyxl.load_workbook(fpath)

# 엑셀 시트 선택
ws = wb['오징어게임']

# 데이터 수정하기
ws['A3'] = 456
ws['B3'] = '성기훈'

# 데이터 저장하기
wb.save(fpath)