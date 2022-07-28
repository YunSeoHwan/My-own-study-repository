import requests
from bs4 import BeautifulSoup
import openpyxl

wb = openpyxl.Workbook()
ws = wb.create_sheet('주식정보')

codes = [
    '005930',
    '000660',
    '035720'
]
i = 1
for code in codes:
    i += 1
    url = f'https://finance.naver.com/item/sise.naver?code={code}'
    response = requests.get(url)
    html = response.text

    soup = BeautifulSoup(html, 'html.parser')
    price = soup.select_one('#_nowVal').text
    price = price.replace(',', '')
    ws[f'B{i}'] = int(price)
wb.save(r'C:\python_crawling\C3\stock_data.xlsx')