import openpyxl

# 엑셀 만들기
wb = openpyxl.Workbook()

# 워크시트 만들기
ws = wb.create_sheet('오징어게임')

# 데이터 추가하기
ws['A1'] = '참가번호'
ws['B1'] = '성명'

ws['A2'] = 1
ws['B2'] = '오일남'

# 엑셀 저장하기
wb.save(r'C:\python_crawling\C3\참가자_data.xlsx')